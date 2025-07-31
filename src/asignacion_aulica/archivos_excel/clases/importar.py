from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from pathlib import Path
import openpyxl

from asignacion_aulica.validación_de_datos.excepciones import DatoInválidoException, ExcelInválidoException
from asignacion_aulica.validación_de_datos.validaciones import *
from asignacion_aulica.archivos_excel.clases.clase import Clase

COLUMNAS = (
    'Año',
    'Materia',
    'Cuatrimestral / Anual',
    'Comisión',
    'Teórica / Práctica',
    'Día',
    'Horario inicio',
    'Horario fin',
    'Cupo',
    'Docente',
    'Auxiliar',
    'Promocionable\nSi (Nota) / No',
    'Lugar',
    'Aula'
)

def _cell_coordinates(cell: Cell) -> str:
    '''
    Dada una celda, devuelve sus coordenadas en un formato legible para el usuario.
    Por ejemplo: F5.
    '''
    return cell.column_letter+str(cell.row)

def _separar_celdas_unidas(hoja: Worksheet, min_row: int):
    '''
    Separa todos los grupos de celdas unidas que haya por debajo de la fila
    `min_row`.

    Luego de esta operación, todas las celdas que previamente estaban unidas
    quedan separadas pero mantienen el valor que tenían.
    '''
    for rango in list(hoja.merged_cells.ranges):
        valor = hoja.cell(rango.min_row, rango.min_col).value
        hoja.unmerge_cells(rango.coord)

        for fila in hoja.iter_rows(rango.min_row, rango.max_row, rango.min_col, rango.max_col):
            for celda in fila:
                celda.value = valor

def leer_preámbulo(hoja: Worksheet) -> tuple[str, int, str]:
    '''
    Verifica que el preámbulo de una hoja del archivo sea válido y extrae sus
    datos.

    :return: Nombre de la carrera, año, cuatrimestre.
    :raise ExcelInválidoException: Si el formato del preámbulo no es válido.
    :raise DatoInválidoException: Si el contenido del preámbulo no es válido.
    '''
    if hoja['C1'].value != 'Carrera: ' or hoja['C2'].value != 'Año: ' or hoja['E2'].value != 'Cuatrimestre: ':
        raise ExcelInválidoException('El encabezado del archivo fue modificado. Por favor no modificar la estructura de la plantilla.')
    
    carrera = validar_str_no_vacío(hoja['D1'].value, 'El nombre de la carrera en la celda D1 no puede estar vacío.')
    año = validar_año(hoja['D2'].value, 'El valor de la celda D2 ')
    cuatrimestre = str_posiblemente_vacío(hoja['G2'].value)

    return carrera, año, cuatrimestre

def leer_tabla(hoja: Worksheet) -> list[Clase]:
    '''
    Lee los datos de clases de una hoja del excel.

    :return: Un iterable de clases. TODO: Determinar en qué formato conviene
        devolver los datos cuando esté la interfaz del módulo de gestión de datos.
    :raise DatoInválidoException: Si hay un dato inválido.
    :raise ExcelInválidoException: Si el formato de la tabla no es correcto.
    '''
    # Verificar que no se hayan modificado las columnas
    if hoja.max_column < len(COLUMNAS):
        raise ExcelInválidoException('Se quitaron columnas de la tabla. Por favor no modificar la estructura de la plantilla.')
    
    fila_títulos = 3
    for índice, título in enumerate(COLUMNAS):
        if hoja.cell(fila_títulos, índice+1).value != título:
            raise ExcelInválidoException(f'Se cambió una columna de la tabla en la celda {get_column_letter(índice+1)}{fila_títulos}. Por favor no modificar la estructura de la plantilla.')
    
    # Si hay celdas unidas dentro de la tabla, separarlas
    _separar_celdas_unidas(hoja, fila_títulos+1)

    # Cargar y verificar datos
    clases = []
    for fila in hoja.iter_rows(fila_títulos+1, hoja.max_row, 1, len(COLUMNAS)):
        fila_no_vacía = any(map(lambda cell: cell.value, fila))
        if fila_no_vacía:
            año =              validar_año_del_plan_de_estudios(fila[ 0].value, f'En la celda {_cell_coordinates(fila[0])}: ')
            materia =                      validar_str_no_vacío(fila[ 1].value, f'En la celda {_cell_coordinates(fila[1])}: el nombre de la materia no debe estar vacío.')
            cuatrimestral_o_anual =      str_posiblemente_vacío(fila[ 2].value)
            comisión =                   str_posiblemente_vacío(fila[ 3].value)
            teórica_o_práctica =         str_posiblemente_vacío(fila[ 4].value)
            día =                                   validar_día(fila[ 5].value, f'En la celda {_cell_coordinates(fila[5])}: ')
            horario_inicio =                   debería_ser_time(fila[ 6].value, f'En la celda {_cell_coordinates(fila[6])}: ')
            horario_fin =                      debería_ser_time(fila[ 7].value, f'En la celda {_cell_coordinates(fila[7])}: ')
            cantidad_de_alumnos = validar_int_positivo_opcional(fila[ 8].value, f'En la celda {_cell_coordinates(fila[8])}: el cupo ')
            docente =                    str_posiblemente_vacío(fila[ 9].value)
            auxiliar =                   str_posiblemente_vacío(fila[10].value)
            promocionable =              str_posiblemente_vacío(fila[11].value)
            edificio =                   str_posiblemente_vacío(fila[12].value)
            aula =                       str_posiblemente_vacío(fila[13].value)
            
            virtual = edificio.lower() == 'virtual'
            if virtual:
                edificio = ''

            if not virtual and cantidad_de_alumnos is None:
                raise DatoInválidoException(f'En la celda {_cell_coordinates(fila[8])}: el cupo de las clases presenciales no debe estar vacío.')
        
            clases.append(Clase(
                año, materia, cuatrimestral_o_anual, comisión, teórica_o_práctica,
                día, horario_inicio, horario_fin, cantidad_de_alumnos, docente,
                auxiliar, promocionable, virtual, edificio, aula
            ))
    
    return clases

def importar(filename: str|Path) -> list[ tuple[str, int, str, list[Clase]] ]:
    '''
    Leer datos de un archivo excel con el formato de la plantilla de clases.

    :param filename: El path absoluto del archivo.

    :return: Una lista de tuplas (carrera, año, cuatrimestre, clases). Cada
        elemento de la lista contiene los datos de una página del archivo.

    :raise FileNotFoundError: Si el archivo no existe.
    :raise openpyxl.utils.exceptions.InvalidFileException: Si el archivo no es un archivo excel.
    :raise ExcelInválidoException: Si el formato del archivo no es correcto.
    :raise DatoInválidoException: Si el archivo contiene un dato inválido.
    '''
    archivo = openpyxl.load_workbook(filename)
    nombres_de_las_hojas = archivo.sheetnames

    data = []

    for nombre in nombres_de_las_hojas:
        hoja = archivo[nombre]
        try:
            carrera, año, cuatrimestre = leer_preámbulo(hoja)
            clases = leer_tabla(hoja)
            data.append((carrera, año, cuatrimestre, clases))
        except (ExcelInválidoException, DatoInválidoException) as e:
            # Tirar la misma excepción, pero agregando el nombre de la hoja al mensaje
            raise type(e)(f'En la hoja {nombre}: ' + str(e))

    return data

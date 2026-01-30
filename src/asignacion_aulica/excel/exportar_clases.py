from itertools import groupby
from collections.abc import Iterable
from typing import Any
from openpyxl.styles import Border, PatternFill
from openpyxl.workbook.workbook import Workbook
from collections.abc import Sequence
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet

from asignacion_aulica.excel.estilos import borde_gris, fill_blanco, fill_rojo_super_clarito
from asignacion_aulica.excel.plantilla_clases import (
    celda_nombre_carrera,
    generar_plantilla,
    fila_primer_clase,
    Columna
)
from asignacion_aulica.gestor_de_datos.entidades import Carrera, Clase, Materia

class RowCounter:
    def __init__(self, initial_value: int = 0):
        self._current: int = initial_value
    def get(self) -> int:
        return self._current
    def next(self) -> int:
        self._current += 1
        return self._current

def exportar_datos_de_clases_a_excel(carreras: Sequence[Carrera], path: Path):
    '''
    Escribir los datos de las clases (incluyendo las aulas asignadas) en un
    archivo excel.

    Cada carrera se exporta en una en una hoja del archivo.

    :param carreras: Las carreras cuyos datos se quiere exportar.
    :param path: El path absoluto del archivo.

    :raise TBD: Si no se puede escribir el archivo en el path dado.
    '''
    excel = Workbook()
    excel.remove(excel.active) # Por defecto hay una hoja vacía que no queremos

    for carrera in carreras:
        hoja = excel.create_sheet(title=carrera.nombre)
        generar_plantilla(hoja)
        _escribir_datos_de_una_carrera(hoja, carrera)

    excel.save(path)

def _escribir_datos_de_una_carrera(hoja: Worksheet, carrera: Carrera):
    hoja[celda_nombre_carrera].value = carrera.nombre
    #TODO: ano y cuatrimestre

    # Copiamos la lista de materias para poder ordenarla sin afectar al gestor.
    # Ordenamos las materias primero por año y después alfabéticamente.
    materias = list(carrera.materias)
    materias.sort(key=lambda materia: (materia.año, materia.nombre))

    fila_actual = RowCounter(initial_value=fila_primer_clase)
    par = True
    for materia in materias:
        if len(materia.clases) > 0:
            _escribir_datos_de_una_materia(
                hoja,
                materia,
                fila_actual,
                # Colores de fondo intercalados para mejor legibilidad
                fill=fill_blanco if par else fill_rojo_super_clarito
            )
            par = not par

def _escribir_datos_de_una_materia(hoja: Worksheet, materia: Materia, fila_actual: RowCounter, fill: PatternFill):
    columnas_a_unir: tuple[tuple[Columna, str|int], ...] = (
        (Columna.año, materia.año),
        (Columna.materia, materia.nombre),
        (Columna.cuatrimestral_o_anual, materia.cuatrimestral_o_anual)
    )
    for columna, valor in columnas_a_unir:
        _merge_cells_and_set_value(
            hoja,
            valor,
            start_row=fila_actual.get(),
            start_col=columna,
            n_rows=len(materia.clases),
            border=Border(bottom=borde_gris, top=borde_gris),
            fill=fill
        )

    # Copiamos la lista de clases para poder ordenarla sin afectar al gestor.
    # Ordenamos las clases primero por comisión, después por día, y después por horario de inicio
    clases = list(materia.clases)
    clases.sort(key=lambda clase: (clase.comisión, clase.día, clase.horario.inicio))
    
    clases_por_comisión = groupby(clases, key=lambda clase: clase.comisión)
    for _, clases_de_la_comisión in clases_por_comisión:
        _escribir_datos_de_una_comisión(hoja, clases_de_la_comisión, fila_actual, fill)

def _escribir_datos_de_una_comisión(hoja: Worksheet, clases: Iterable[Clase], fila_actual: RowCounter, fill: PatternFill):
    primera_fila_de_la_comisión = fila_actual.get()

    for clase in clases:
        _escribir_datos_de_una_clase(hoja, clase, fila_actual.get(), fill)
        fila_actual.next()
    
    # Unir las celdas que tienen los mismos valores
    columnas_a_unir = tuple(
        col for col in Columna
        if col not in (Columna.año, Columna.materia, Columna.cuatrimestral_o_anual)
    )
    for col in columnas_a_unir:
        _merge_vertically_neighboring_cells_with_equal_value(
            hoja, col, primera_fila_de_la_comisión, fila_actual.get()-primera_fila_de_la_comisión
        )
        hoja.cell(primera_fila_de_la_comisión, col).border=Border(top=borde_gris)
        hoja.cell(fila_actual.get()-1, col).border=Border(bottom=borde_gris)

def _escribir_datos_de_una_clase(hoja: Worksheet, clase: Clase, fila: int, fill: PatternFill):
    columnas_y_valores: tuple[tuple[int, Any], ...] = (
        (Columna.comisión, clase.comisión),
        (Columna.teórica_o_práctica, clase.teórica_o_práctica),
        (Columna.día, clase.día.name),
        (Columna.horario_inicio, clase.horario.inicio),
        (Columna.horario_fin, clase.horario.fin),
        (Columna.cupo, clase.cantidad_de_alumnos),
        (Columna.docente, clase.docente),
        (Columna.auxiliar, clase.auxiliar),
        (Columna.promocionable, clase.promocionable),
        (Columna.lugar,
            'Virtual'
            if clase.virtual
            else clase.aula_asignada.edificio.nombre + ' - ' + clase.aula_asignada.nombre
            if clase.aula_asignada is not None
            else ''
        )
    )
    for columna, valor in columnas_y_valores:
        cell = hoja.cell(fila, columna)
        cell.value = valor
        cell.fill = fill

def _merge_cells_and_set_value(
    sheet: Worksheet,
    value: Any,
    start_row: int,
    start_col: int,
    n_rows: int = 1,
    n_cols: int = 1,
    border: Border|None = None,
    fill: PatternFill|None = None
):
    sheet.merge_cells(
        start_row=start_row,
        end_row=start_row + n_rows - 1,
        start_column=start_col,
        end_column=start_col + n_cols - 1
    )
    merged_cell = sheet.cell(start_row, start_col)
    merged_cell.value = value
    if fill:
        merged_cell.fill = fill
    if border:
        merged_cell.border = border

def _merge_vertically_neighboring_cells_with_equal_value(
    sheet: Worksheet,
    column: int,
    start_row: int,
    n_rows: int,
):
    end_row = start_row + n_rows - 1
    start_of_range = start_row
    while start_of_range < end_row:
        end_of_range = _find_rows_with_equal_values(sheet, column, start_of_range, end_row)
        if end_of_range > start_of_range:
            sheet.merge_cells(start_row=start_of_range, end_row=end_of_range, start_column=column, end_column=column)
        start_of_range = end_of_range + 1

def _find_rows_with_equal_values(
    sheet: Worksheet,
    column: int,
    start_row: int,
    max_row: int
):
    value = sheet.cell(row=start_row, column=column).value
    current_row = start_row
    while current_row < max_row and sheet.cell(row=current_row+1, column=column).value == value:
        current_row += 1
    
    return current_row

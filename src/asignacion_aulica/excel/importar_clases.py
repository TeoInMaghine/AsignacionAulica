from dataclasses import dataclass

from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
import openpyxl

from asignacion_aulica.excel.plantilla_clases import CeldaEncabezado, Columna, TÍTULOS_DE_COLUMNAS, fila_header, fila_primer_clase, n_columnas
from asignacion_aulica.gestor_de_datos.días_y_horarios import RangoHorario
from asignacion_aulica.validación_de_datos.excepciones import DatoInválidoException, ExcelInválidoException
from asignacion_aulica.validación_de_datos.validaciones import *

@dataclass
class ClaseLeída:
    día: Día
    horario: RangoHorario
    virtual: bool
    cantidad_de_alumnos: int|None
    edificio: str|None
    aula: str|None
    comisión: str
    teórica_o_práctica: str
    promocionable: str
    docente: str
    auxiliar: str

@dataclass
class MateriaLeída:
    año: int
    nombre: str
    cuatrimestral_o_anual: str
    clases: list[ClaseLeída]

@dataclass
class CarreraLeída:
    nombre: str
    año: int
    cuatrimestre: str
    materias: list[MateriaLeída]

def importar_clases_de_excel(filename: str|Path) -> list[CarreraLeída]:
    '''
    Leer datos de un archivo excel con el formato de la plantilla de clases.

    :param filename: El path absoluto del archivo.

    :return: Una lista con los datos leídos de cada página del archivo.

    :raise FileNotFoundError: Si el archivo no existe.
    :raise openpyxl.utils.exceptions.InvalidFileException: Si el archivo no es un archivo excel.
    :raise ExcelInválidoException: Si el formato del archivo no es correcto.
    :raise DatoInválidoException: Si el archivo contiene un dato inválido.
    '''
    archivo = openpyxl.load_workbook(filename)
    if len(archivo.sheetnames) == 0:
        raise ExcelInválidoException('El archivo debe tener al menos una hoja.')

    nombres_de_las_hojas = archivo.sheetnames
    data: list[CarreraLeída] = []

    for nombre in nombres_de_las_hojas:
        hoja = archivo[nombre]
        try:
            carrera, año, cuatrimestre = leer_encabezado(hoja)
            materias = leer_materias(hoja)
            data.append(CarreraLeída(carrera, año, cuatrimestre, materias))
        except (ExcelInválidoException, DatoInválidoException) as e:
            # Tirar la misma excepción, pero agregando el nombre de la hoja al mensaje
            raise type(e)(f'En la hoja {nombre}: ' + str(e))

    return data

def leer_encabezado(hoja: Worksheet) -> tuple[str, int, str]:
    '''
    Verifica que el encabezado de una hoja del archivo sea válido y extrae sus
    datos.

    :return: Nombre de la carrera, año, cuatrimestre.
    :raise ExcelInválidoException: Si el formato del preámbulo no es válido.
    :raise DatoInválidoException: Si el contenido del preámbulo no es válido.
    '''
    if (
        hoja[CeldaEncabezado.label_carrera].value != 'Carrera: '
        or hoja[CeldaEncabezado.label_año].value != 'Año: '
        or hoja[CeldaEncabezado.label_cuatrimestre].value != 'Cuatrimestre: '
    ):
        raise ExcelInválidoException('El encabezado del archivo fue modificado. Por favor no modificar la estructura de la plantilla.')
    
    carrera = validar_str_no_vacío(
        hoja[CeldaEncabezado.carrera].value,
        'El nombre de la carrera en la celda D1 no puede estar vacío.'
    )
    año = validar_año(
        hoja[CeldaEncabezado.año].value,
        'El valor de la celda D2 '
    )
    cuatrimestre = str_posiblemente_vacío(hoja[CeldaEncabezado.cuatrimestre].value)

    return carrera, año, cuatrimestre

def leer_materias(hoja: Worksheet) -> list[MateriaLeída]:
    '''
    Lee los datos de materias y clases de una hoja del excel.

    :raise DatoInválidoException: Si hay un dato inválido.
    :raise ExcelInválidoException: Si el formato de la tabla no es correcto.
    '''
    # Verificar que no se hayan modificado las columnas
    if hoja.max_column < n_columnas:
        raise ExcelInválidoException('Se quitaron columnas de la tabla. Por favor no modificar la estructura de la plantilla.')
    
    for columna, título in zip(Columna, TÍTULOS_DE_COLUMNAS):
        if hoja.cell(fila_header, columna).value != título:
            raise ExcelInválidoException(f'Se cambió una columna de la tabla en la celda {get_column_letter(columna)}{fila_header}. Por favor no modificar la estructura de la plantilla.')
    
    # Si hay celdas unidas dentro de la tabla, separarlas
    _separar_celdas_unidas(hoja, fila_header)

    # Cargar y verificar datos
    materias: dict[str, MateriaLeída] = {}
    for fila in hoja.iter_rows(fila_primer_clase, hoja.max_row, 1, n_columnas):
        fila_vacía = all(map(lambda cell: not cell.value, fila))
        if fila_vacía: continue

        nombre_materia: str = validar_str_no_vacío(
            fila[Columna.materia-1].value,
            f'En la celda {_cell_coordinates(fila[Columna.materia-1])}: el nombre de la materia no debe estar vacío.'
        )
        if nombre_materia not in materias:
            año = validar_año_del_plan_de_estudios(fila[Columna.año-1].value, f'En la celda {_cell_coordinates(fila[Columna.año-1])}: ')
            cuatrimestral_o_anual = str_posiblemente_vacío(fila[Columna.cuatrimestral_o_anual-1].value)
            materias[nombre_materia] = MateriaLeída(año, nombre_materia, cuatrimestral_o_anual, clases=[])
        
        materias[nombre_materia].clases.append(_leer_clase(fila))
    
    if len(materias) == 0:
        raise DatoInválidoException('La carrera debe tener al menos una clase.')
    
    return list(materias.values())

def _leer_clase(fila: tuple[Cell, ...]) -> ClaseLeída:
    comisión = str_posiblemente_vacío(fila[Columna.comisión-1].value)
    cantidad_de_alumnos = validar_int_positivo_opcional(fila[Columna.cupo-1].value, f'En la celda {_cell_coordinates(fila[Columna.cupo-1])}: el cupo ')
    día = validar_día(fila[Columna.día-1].value, f'En la celda {_cell_coordinates(fila[Columna.día-1])}: ')
    horario_inicio = debería_ser_time(fila[Columna.horario_inicio-1].value, f'En la celda {_cell_coordinates(fila[Columna.horario_inicio-1])}: ')
    horario_fin = debería_ser_time(fila[Columna.horario_fin-1].value, f'En la celda {_cell_coordinates(fila[Columna.horario_fin-1])}: ')
    lugar = str_posiblemente_vacío(fila[Columna.lugar-1].value)
    teórica_o_práctica = str_posiblemente_vacío(fila[Columna.teórica_o_práctica-1].value)
    docente = str_posiblemente_vacío(fila[Columna.docente-1].value)
    auxiliar = str_posiblemente_vacío(fila[Columna.auxiliar-1].value)
    promocionable = str_posiblemente_vacío(fila[Columna.promocionable-1].value)

    if len(lugar) == 0:
        virtual = False
        edificio = None
        aula = None
    elif lugar.lower() == 'virtual':
        virtual = True
        edificio = None
        aula = None
    elif lugar.count(' - ') != 1:
        raise DatoInválidoException(f'En la celda {_cell_coordinates(fila[Columna.lugar-1])}: "{lugar}" no se reconoce como un nombre de edificio y aula.')
    else:
        virtual = False
        edificio, aula = lugar.split(' - ')
    
    if not virtual and cantidad_de_alumnos is None:
        raise DatoInválidoException(f'En la celda {_cell_coordinates(fila[Columna.cupo-1])}: el cupo de las clases presenciales no debe estar vacío.')

    return ClaseLeída(
        día, RangoHorario(horario_inicio, horario_fin), virtual,
        cantidad_de_alumnos, edificio, aula, comisión, teórica_o_práctica,
        promocionable, docente, auxiliar
    )
    
    

def _separar_celdas_unidas(hoja: Worksheet, min_row: int):
    '''
    Separa todos los grupos de celdas unidas que haya por debajo de la fila
    `min_row`.

    Luego de esta operación, todas las celdas que previamente estaban unidas
    quedan separadas pero mantienen el valor que tenían.
    '''
    for rango in list(hoja.merged_cells.ranges):
        if rango.min_row < min_row: continue

        valor = hoja.cell(rango.min_row, rango.min_col).value
        hoja.unmerge_cells(rango.coord)

        for fila in hoja.iter_rows(rango.min_row, rango.max_row, rango.min_col, rango.max_col):
            for celda in fila:
                celda.value = valor

def _cell_coordinates(cell: Cell) -> str:
    '''
    Dada una celda, devuelve sus coordenadas en un formato legible para el usuario.
    Por ejemplo: F5.
    '''
    return cell.column_letter+str(cell.row)

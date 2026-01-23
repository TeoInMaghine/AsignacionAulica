'''
Este modulo define la plantilla para el archivo excel de clases/horarios/aulas.

La plantilla está basada en el archivo excel que les directores de carreras le
pasan al encargado de la asignación. Se modificaron algunas cosas para que sea
más usable y automatizable.

La plantilla tiene un encabezado que consiste de:
- El logo de la universidad
- El nombre de la carrera
- El año y número de cuatrimestre

Abajo del encabezado hay una tabla con varias columnas de datos. La tabla tiene
una fila por cada clase, posiblemente con celdas unidas verticalmente (por
ejemplo, la columna "Materia" puede estar unida en las filas de las clases de
cada materia).
'''
from collections.abc import Sequence
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles.borders import Border
from openpyxl import Workbook

import sys

from asignacion_aulica.excel.estilos import (
    get_logo,
    estilo_header,
    estilo_horarios,
    estilo_tabla,
    fill_rojo_unrn,
    font_default,
    font_encabezado,
    font_bold,
    a_la_derecha,
    a_la_izquierda,
    borde_blanco,
    borde_negro
)

from asignacion_aulica.excel.validadores import (
    no_cambiar_este_valor,
    año_del_calendario,
    año_del_plan_de_estudios,
    número_natural,
    día_de_la_semana,
    horario
)

COLUMNAS: Sequence[str] = (
    'Año',
    'Materia',
    'Cuatrimestral / Anual',
    'Comisión',
    'Teórica / Práctica',
    'Día',
    'Horario inicio',
    'Horario fin',
    'Cupo',
    'Docente',
    'Auxiliar',
    'Promocionable\nSi (Nota) / No',
    'Lugar',
    'Aula'
)

n_columnas = len(COLUMNAS)
última_columna: str = get_column_letter(n_columnas)

# Cantidad de filas a las que aplicar formato
n_filas = 200

def generar_encabezado(hoja: Worksheet):
    # Configurar tamaño de las filas
    altura_filas = font_encabezado.size + 7
    hoja.row_dimensions[1].height = altura_filas
    hoja.row_dimensions[2].height = altura_filas

    # Marcar bordes externos
    hoja.column_dimensions['A'].border = Border(left=borde_negro)
    hoja.column_dimensions[última_columna].border = Border(right=borde_negro)

    # Insertar el logo
    logo = get_logo(2*altura_filas)
    hoja.add_image(logo, 'A1')
    hoja.merge_cells(start_row=1, end_row=2, start_column=1, end_column=4)
    no_cambiar_este_valor.add('A1:D2')
    hoja.cell(1, 1).fill = fill_rojo_unrn
    hoja.cell(1, 1).border = Border(top=borde_negro, left=borde_negro)
    hoja.cell(1, 2).border = Border(top=borde_negro)
    hoja.cell(1, 3).border = Border(top=borde_negro)
    hoja.cell(1, 4).border = Border(top=borde_negro)

    hoja.cell(2, 1).border = Border(bottom=borde_negro, left=borde_negro)
    hoja.cell(2, 2).border = Border(bottom=borde_negro)
    hoja.cell(2, 3).border = Border(bottom=borde_negro)
    hoja.cell(2, 4).border = Border(bottom=borde_negro)

    # Fila con el nombre de la carrera
    cell = hoja.cell(1, 5, value='Carrera: ')
    no_cambiar_este_valor.add(cell)
    cell.fill = fill_rojo_unrn
    cell.font = font_encabezado
    cell.alignment = a_la_derecha
    cell.border = Border(top=borde_negro, bottom=borde_blanco)

    hoja.merge_cells(start_row=1, end_row=1, start_column=6, end_column=n_columnas)
    cell = hoja.cell(1, 6, value='') # Celda para completar el nombre de la carrera
    cell.fill = fill_rojo_unrn
    cell.font = font_encabezado
    cell.alignment = a_la_izquierda

    for i in range(6, n_columnas):
        hoja.cell(1, i).border = Border(top=borde_negro, bottom=borde_blanco)
    hoja.cell(1, n_columnas).border = Border(top=borde_negro, bottom=borde_blanco, right=borde_negro)

    # Fila con el año y cuatrimestre
    cell = hoja.cell(2, 5, value='Año: ')
    no_cambiar_este_valor.add(cell)
    cell.fill = fill_rojo_unrn
    cell.font = font_encabezado
    cell.alignment = a_la_derecha

    cell = hoja.cell(2, 6, value='') # Celda para completar el año
    año_del_calendario.add(cell)
    cell.fill = fill_rojo_unrn
    cell.font = font_encabezado
    cell.alignment = a_la_izquierda

    hoja.merge_cells(start_row=2, end_row=2, start_column=7, end_column=9)
    no_cambiar_este_valor.add(f'{get_column_letter(7)}2:{get_column_letter(9)}2')
    cell = hoja.cell(2, 7, value='Cuatrimestre: ')
    cell.fill = fill_rojo_unrn
    cell.font = font_encabezado
    cell.alignment = a_la_derecha

    hoja.merge_cells(start_row=2, end_row=2, start_column=10, end_column=n_columnas)
    cell = hoja.cell(2, 10, value='') # Celda para completar el cuatrimestre
    cell.fill = fill_rojo_unrn
    cell.font = font_encabezado
    cell.alignment = a_la_izquierda

def generar_tabla(hoja: Worksheet):
    # Insertar fila con los nombres de las columnas
    hoja.append(COLUMNAS)
    fila_header = hoja.max_row
    no_cambiar_este_valor.add(f'A{fila_header}:{última_columna}{fila_header}')

    # Bloquear movimiento de los nombres para que se mantangan visibles al escrollear
    hoja.freeze_panes = hoja.cell(fila_header+1, 1)

    # Configurar estilo de los nombres
    for i in range(1, n_columnas+1):
        hoja.cell(fila_header, i).style = estilo_header
    
    # Ajustar tamaños de las columnas (los números fueron calibrados por prueba y error)
    font_size_ratio = estilo_header.font.size / 11
    hoja.column_dimensions['A'].width =  5 * font_size_ratio # Año
    hoja.column_dimensions['B'].width = 25 * font_size_ratio # Materia
    hoja.column_dimensions['C'].width = 12 * font_size_ratio # Cuatrimestral o anual
    hoja.column_dimensions['D'].width = 10 * font_size_ratio # Comisión
    hoja.column_dimensions['E'].width = 11 * font_size_ratio # Teórica o práctica
    hoja.column_dimensions['F'].width = 11 * font_size_ratio # Día
    hoja.column_dimensions['G'].width =  7 * font_size_ratio # Horario de inicio
    hoja.column_dimensions['H'].width =  7 * font_size_ratio # Horario de fin
    hoja.column_dimensions['I'].width =  5 * font_size_ratio # Cupo
    hoja.column_dimensions['J'].width = 12 * font_size_ratio # Docente
    hoja.column_dimensions['K'].width = 12 * font_size_ratio # Auxiliar
    hoja.column_dimensions['L'].width = 14 * font_size_ratio # Promocionable
    hoja.column_dimensions['M'].width = 12 * font_size_ratio # Lugar
    hoja.column_dimensions['N'].width =  8 * font_size_ratio # Aula

    # Agregar validadores
    año_del_plan_de_estudios.add(f'A{fila_header+1}:A1048576') # 1048576 significa hasta el final de la columna.
    día_de_la_semana.add(f'F{fila_header+1}:F1048576') # Día
    horario.add(f'G{fila_header+1}:G1048576') # Horario de inicio
    horario.add(f'H{fila_header+1}:H1048576') # Horario de fin
    número_natural.add(f'I{fila_header+1}:I1048576') # Cupo

    # Configurar estilo de la tabla hasta la fila n_filas
    columnas_horario = (7, 8)
    for row in range(fila_header + 1, fila_header + n_filas):
        for col in range(1, n_columnas):
            if col in columnas_horario:
                hoja.cell(row, col).style = estilo_horarios
            else:
                hoja.cell(row, col).style = estilo_tabla

def crear_plantilla() -> Workbook:
    plantilla = Workbook()
    plantilla._fonts[0] = font_default

    hoja = plantilla.active
    hoja.title = 'Plantilla'
    hoja.add_data_validation(no_cambiar_este_valor)
    hoja.add_data_validation(año_del_plan_de_estudios)
    hoja.add_data_validation(año_del_calendario)
    hoja.add_data_validation(número_natural)
    hoja.add_data_validation(día_de_la_semana)
    hoja.add_data_validation(horario)
    
    generar_encabezado(hoja)
    generar_tabla(hoja)

    return plantilla

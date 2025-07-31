from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Fill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.units import points_to_pixels
from openpyxl.drawing.image import Image
from copy import copy
from os import path

este_directorio = path.split(__file__)[0]

logo_path = path.join(este_directorio, 'unrn_logo.png')

rojo_unrn = 'EB2242'
fill_rojo_unrn = PatternFill(patternType='solid', fgColor=rojo_unrn)

font_default = Font(name = 'arial', size = 12)

font_bold = copy(font_default)
font_bold.bold = True

font_preámbulo = copy(font_bold)
font_preámbulo.size = 18
font_preámbulo.color = 'FFFFFF'

centrado = Alignment(horizontal='center', vertical='center', wrap_text=True)
a_la_derecha = Alignment(horizontal = 'right', vertical='center')
a_la_izquierda = Alignment(horizontal = 'left', vertical='center')

borde_negro = Side(border_style='thin', color='000000')
borde_negro_grueso = Side(border_style='medium', color='000000')
borde_blanco = Side(border_style='thin', color='FFFFFF')
todos_los_bordes_negros = Border(top=borde_negro, bottom=borde_negro, left=borde_negro, right=borde_negro)

def get_logo(height_points):
    imagen = Image(logo_path)
    scale_ratio = points_to_pixels(height_points) / imagen.height
    imagen.height *= scale_ratio
    imagen.width *= scale_ratio

    return imagen

def format_column_as_24hs_time(
    sheet: Worksheet,
    column_index: int,
    start_row: int,
    end_row: int
    ):
    for i in range(start_row, end_row):
        sheet.cell(i, column_index).number_format = 'HH:MM'

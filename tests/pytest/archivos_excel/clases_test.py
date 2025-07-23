import openpyxl
import pytest

from asignacion_aulica.archivos_excel.clases.importar import leer_preámbulo
from asignacion_aulica.validación_de_datos.excepciones import DatoInválidoException, ExcelInválidoException

@pytest.mark.archivo('clases_nominal.xlsx')
def test_preámbulo_nominal(archivo):
    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active

    carrera, año, cuatrimestre = leer_preámbulo(hoja)
    assert carrera == 'Acá va el nombre de la carrera'
    assert año == 2025
    assert cuatrimestre == 'Primero'

@pytest.mark.archivo('clases_preámbulo_roto.xlsx')
def test_preámbulo_roto(archivo):
    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active

    with pytest.raises(ExcelInválidoException):
        leer_preámbulo(hoja)

@pytest.mark.archivo('clases_carrera_vacía.xlsx')
def test_carrera_vacía(archivo):
    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active

    with pytest.raises(DatoInválidoException):
        leer_preámbulo(hoja)


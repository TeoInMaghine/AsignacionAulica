from datetime import time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from asignacion_aulica.excel.exportar_clases import exportar_datos_de_clases_a_excel
from asignacion_aulica.excel.plantilla_clases import fila_primer_clase, Columna
from asignacion_aulica.gestor_de_datos.días_y_horarios import RangoHorario, Día
from asignacion_aulica.gestor_de_datos.entidades import Carreras
from mocks import MockCarrera, MockMateria, MockClase, MockEdificio, MockAula
import pytest

@pytest.fixture
def tmp_xlsx_filename(tmp_path: Path) -> Path:
    '''
    :return: A filename inside a temporary directory created for each  test.
    '''
    return tmp_path / "tmp_file.xlsx"

@pytest.fixture
def excel_exportado(carreras: Carreras, tmp_xlsx_filename: Path) -> Workbook:
    '''
    Exporta los datos del fixture de carreras a un archivo excel, para luego
    leer y devolver el contenido del archivo.
    '''
    exportar_datos_de_clases_a_excel(carreras, tmp_xlsx_filename)
    return openpyxl.load_workbook(tmp_xlsx_filename)

def get_cell_value(sheet: Worksheet, row: int, col: int) -> Any:
    '''
    :return: The value of a cell that may be merged or not.

    For empy cells, return an empty string instead of `None`.
    
    https://stackoverflow.com/a/76617485
    '''
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # return the left top cell
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return '' if cell.value is None else cell.value

@pytest.mark.carreras(
    MockCarrera(nombre='A'),
    MockCarrera(nombre='B'),
    MockCarrera(nombre='C')
)
def test_encabezado(carreras: Carreras, excel_exportado: Workbook):
    '''
    Verificar que hay una hoja para cada carrera y que aparecen los datos
    correctos en el encabezado.
    '''
    assert len(excel_exportado.worksheets) == len(carreras)
    for hoja, carrera in zip(excel_exportado.worksheets, carreras):
        assert hoja.title == carrera.nombre
        assert hoja['F1'].value == carrera.nombre
        #TODO: año y cuatrimestre
    
@pytest.mark.edificios(
    MockEdificio(nombre='E1', aulas=(MockAula(nombre='A1.1'), MockAula(nombre='A1.2'))),
    MockEdificio(nombre='E2', aulas=(MockAula(nombre='A2.1'), MockAula(nombre='A2.2')))
)
@pytest.mark.carreras(
    MockCarrera(
        materias=(
            MockMateria(
                nombre='Ma 1',
                año = 8,
                cuatrimestral_o_anual = 'no sé',
                clases=(
                    MockClase(
                        teórica_o_práctica='si',
                        aula_asignada=None,
                        docente='Uno re copado'
                    ),
                    MockClase(
                        día=Día.Jueves,
                        cantidad_de_alumnos=69,
                        aula_asignada=(1, 0)
                    ),
                    MockClase(
                        comisión='otra',
                        cantidad_de_alumnos=2,
                        auxiliar='meh',
                        virtual=True
                    ),
                    MockClase(
                        comisión='una',
                        horario=RangoHorario(time(7,15), time(8,45)),
                        promocionable='Si (10)',
                        aula_asignada=(0, 1)
                    ),
                )
            ),
        )
    )
)
def test_datos_de_las_clases(carreras: Carreras, excel_exportado: Workbook):
    carrera = carreras[0]
    materia = carrera.materias[0]
    clases = materia.clases
    
    hoja = excel_exportado.worksheets[0]
    
    for i in range(len(clases)):
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.año) == materia.año
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.materia) == materia.nombre
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.cuatrimestral_o_anual) == materia.cuatrimestral_o_anual
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.comisión) == clases[i].comisión
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.teórica_o_práctica) == clases[i].teórica_o_práctica
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.día) == clases[i].día.name
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.horario_inicio) == clases[i].horario.inicio
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.horario_fin) == clases[i].horario.fin
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.cupo) == clases[i].cantidad_de_alumnos
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.docente) == clases[i].docente
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.auxiliar) == clases[i].auxiliar
        assert get_cell_value(hoja, fila_primer_clase+i, Columna.promocionable) == clases[i].promocionable

    assert get_cell_value(hoja, fila_primer_clase, Columna.edificio) == ''
    assert get_cell_value(hoja, fila_primer_clase, Columna.aula) == ''

    assert get_cell_value(hoja, fila_primer_clase+1, Columna.edificio) == clases[1].aula_asignada.edificio.nombre
    assert get_cell_value(hoja, fila_primer_clase+1, Columna.aula) == clases[1].aula_asignada.nombre

    assert get_cell_value(hoja, fila_primer_clase+2, Columna.edificio) == 'Virtual'
    assert get_cell_value(hoja, fila_primer_clase+2, Columna.aula) == 'Virtual'
    
    assert get_cell_value(hoja, fila_primer_clase+3, Columna.edificio) == clases[3].aula_asignada.edificio.nombre
    assert get_cell_value(hoja, fila_primer_clase+3, Columna.aula) == clases[3].aula_asignada.nombre

@pytest.mark.carreras(
    MockCarrera(
        materias=(
            MockMateria(año = 5, nombre = '5', clases=(MockClase(),)),
            MockMateria(año = 1, nombre = '1', clases=(MockClase(),)),
            MockMateria(año = 2, nombre = '2', clases=(MockClase(),)),
            MockMateria(año = 4, nombre = '4b', clases=(MockClase(),)),
            MockMateria(año = 4, nombre = '4a', clases=(MockClase(),))
        )
    )
)
def test_materias_ordenadas_por_año_y_nombre(excel_exportado: Workbook):
    hoja = excel_exportado.worksheets[0]
    assert get_cell_value(hoja, fila_primer_clase+0, Columna.año) == 1
    assert get_cell_value(hoja, fila_primer_clase+0, Columna.materia) == '1'
    assert get_cell_value(hoja, fila_primer_clase+1, Columna.año) == 2
    assert get_cell_value(hoja, fila_primer_clase+1, Columna.materia) == '2'
    assert get_cell_value(hoja, fila_primer_clase+2, Columna.año) == 4
    assert get_cell_value(hoja, fila_primer_clase+2, Columna.materia) == '4a'
    assert get_cell_value(hoja, fila_primer_clase+3, Columna.año) == 4
    assert get_cell_value(hoja, fila_primer_clase+3, Columna.materia) == '4b'
    assert get_cell_value(hoja, fila_primer_clase+4, Columna.año) == 5
    assert get_cell_value(hoja, fila_primer_clase+4, Columna.materia) == '5'

@pytest.mark.clases(
    MockClase(comisión='b'),
    MockClase(comisión='c', día=Día.Jueves,   horario=RangoHorario(time(12), time(15))),
    MockClase(comisión='a', día=Día.Viernes),
    MockClase(comisión='c', día=Día.Jueves,   horario=RangoHorario(time(10), time(11))),
    MockClase(comisión='a', día=Día.Lunes),
)
def test_clases_ordenadas_por_comisión_día_y_horario(excel_exportado: Workbook):
    hoja = excel_exportado.worksheets[0]

    # La comisión a tiene que quedar primera, con la clase del lunes primero
    assert get_cell_value(hoja, fila_primer_clase+0, Columna.comisión) == 'a'
    assert get_cell_value(hoja, fila_primer_clase+0, Columna.día) == Día.Lunes.name
    assert get_cell_value(hoja, fila_primer_clase+1, Columna.comisión) == 'a'
    assert get_cell_value(hoja, fila_primer_clase+1, Columna.día) == Día.Viernes.name

    # La comisión b tiene que quedar segunda
    assert get_cell_value(hoja, fila_primer_clase+2, Columna.comisión) == 'b'

    # La comisión c tiene que quedar tercera, con la clase más temprano primero
    assert get_cell_value(hoja, fila_primer_clase+3, Columna.comisión) == 'c'
    assert get_cell_value(hoja, fila_primer_clase+3, Columna.horario_inicio) == time(10)
    assert get_cell_value(hoja, fila_primer_clase+4, Columna.comisión) == 'c'
    assert get_cell_value(hoja, fila_primer_clase+4, Columna.horario_inicio) == time(12)
